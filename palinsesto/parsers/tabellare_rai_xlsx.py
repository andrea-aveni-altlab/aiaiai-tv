"""Parser dei listini TABELLARE VIDEO Rai Pubblicità (xlsx, foglio TV-TABELLARE).
Column-driven dagli header (r3-r6), mai offset fissi: i file hanno 2-9
sottoperiodi e larghezze diverse. Output: CSV lungo (rubrica x sottoperiodo)
+ previsione nel DB palinsesto (sorgente='rai_listino')."""
import csv
import os
import re
import sys
from datetime import date, time

import openpyxl

if __package__:
    from .. import db, previsioni
else:
    sys.path.insert(0, str(__import__("pathlib").Path(__file__).parents[2]))
    from palinsesto import db, previsioni

BASE = os.path.expanduser("~/Antigravity/palinsesti_pdf/listini_raipub")
OUT = os.path.join(BASE, "palinsesto_listini.csv")
FILES = [
    "2025/excel/listino_tabellare_video_inverno-2025_aggiornamento-12-dicembre-2024.xlsx",
    "2025/excel/listino_tabellare_video_primavera2025_aggiornamento18luglio2025.xlsx",
    "2025/excel/listino_tabellare_video_estate2025_aggiornamento18luglio2025.xlsx",
    "2025/excel/listino_tabellare_video__settembre_ottobre_2025__aggiornamento18luglio2025.xlsx",
    "2025/excel/listino_novembre-dicembre-2025_tabellare-video_aggiornamento-6novembre2025.xlsx",
    "2025/excel/listino_feste2025-gennaio2026_tabellare_video_aggiorn_22-10-25.xlsx",
    "2026/excel/listino_tabellare_video_inverno2026_aggiornamento-8-gennaio-2026 (1).xlsx",
    "2026/excel/listino_tabellare_video_primavera2026.xlsx",
    "2026/excel/listino_estate-2026_tabellare-video_aggiornamento-24aprile2026.xlsx",
    "2026/excel/listino_tabellare_video_settembre_ottobre_2026.xlsx",
]
GIORNI = ["DOMENICA", "LUNEDI", "MARTEDI", "MERCOLEDI", "GIOVEDI", "VENERDI", "SABATO"]
MESI = {m: i + 1 for i, m in enumerate(
    ["gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno", "luglio",
     "agosto", "settembre", "ottobre", "novembre", "dicembre"])}
TARGET_MAP = {"IND": "individui", "R.A.": "ra", "15-64 ANNI": "15_64",
              "25-54 ANNI": "25_54", "15-34 ANNI": "15_34"}
RE_SOTTOP = re.compile(r"^(\d{1,2})/(\d{1,2})\s*-\s*(\d{1,2})/(\d{1,2})$")
RETE_MAP = {"RAI 1": "RAI1", "RAI 2": "RAI2", "RAI 3": "RAI3", "RAI1": "RAI1",
            "RAI2": "RAI2", "RAI3": "RAI3"}


def pubblicato_da_nome(nome: str):
    t = nome.lower().replace("_", "-")
    m = re.search(r"aggiorn[a-z]*[-\s]?(\d{1,2})[-\s]?([a-z]+)[-\s]?(\d{4})", t)
    if m and m.group(2) in MESI:
        return date(int(m.group(3)), MESI[m.group(2)], int(m.group(1))), "stampata"
    m = re.search(r"aggiorn[a-z]*[-\s]?(\d{1,2})-(\d{1,2})-(\d{2,4})", t)
    if m:
        a = int(m.group(3))
        return date(a + 2000 if a < 100 else a, int(m.group(2)), int(m.group(1))), "stampata"
    return None, None


def orario_split(v):
    if v is None:
        return "", ""
    if isinstance(v, time):
        return v.strftime("%H:%M"), ""
    s = str(v).strip()
    parti = s.split("/")
    def norm(x):
        x = x.strip()
        m = re.match(r"^(\d{1,2})[:.](\d{2})", x)
        return f"{int(m.group(1)):02d}:{m.group(2)}" if m else x
    if len(parti) >= 2:
        return norm(parti[0]), norm(parti[1])
    return norm(parti[0]), ""


def anni_sottoperiodi(labels, anno_base):
    """Assegna gli anni in sequenza: wrap +1 quando il mese decresce."""
    out, anno, prev_m = [], anno_base, None
    for lab in labels:
        m = RE_SOTTOP.match(lab)
        g1, m1, g2, m2 = (int(x) for x in m.groups())
        if prev_m is not None and m1 < prev_m - 6:
            anno += 1
        d1 = date(anno, m1, g1)
        a2 = anno + 1 if m2 < m1 - 6 else anno
        d2 = date(a2, m2, g2)
        out.append((d1, d2))
        prev_m = m2
        anno = a2
    return out


if len(sys.argv) > 1:                      # file espliciti da riga di comando
    FILES = sys.argv[1:]
    BASE = ""
righe_csv, riepilogo = [], []
conn = db.connect()

for rel in FILES:
    path = os.path.join(BASE, rel)
    nome = os.path.basename(rel)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    foglio = next(s for s in wb.sheetnames if s.strip().upper() == "TV-TABELLARE")
    ws = wb[foglio]
    righe = list(ws.iter_rows(min_row=1, max_row=ws.max_row,
                              max_col=ws.max_column, values_only=True))
    periodo_lab = str(righe[1][0] or "").replace("Listino", "").strip()
    anno_base = int(re.search(r"(20\d\d)", periodo_lab + nome).group(1))
    r4, r5, r6 = righe[3], righe[4], righe[5]

    # colonne giorno
    col_giorni = {}
    for j, v in enumerate(r4):
        if v and str(v).strip().upper().replace("'", "") in GIORNI:
            col_giorni[GIORNI.index(str(v).strip().upper())] = j

    # zona tariffe: da r3 'Tariffe' in poi; gruppi in r4, sottoperiodi in r5
    r3 = righe[2]
    col_tariffe_start = next((j for j, v in enumerate(r3)
                              if v and "tariff" in str(v).lower()), None)
    # blocchi stime: etichette d/m-d/m in r4 PRIMA della zona tariffe
    blocchi = []
    for j, v in enumerate(r4):
        if v and RE_SOTTOP.match(str(v).strip()):
            if col_tariffe_start is None or j < col_tariffe_start:
                blocchi.append((j, str(v).strip()))
    date_blocchi = anni_sottoperiodi([b[1] for b in blocchi], anno_base)
    # target di ciascun blocco: da r5, dalla colonna del blocco alla successiva
    limiti = [b[0] for b in blocchi] + [col_tariffe_start or len(r4)]
    blocchi_target = []
    for k, (j0, lab) in enumerate(blocchi):
        tgt = {}
        for j in range(j0, limiti[k + 1]):
            t = str(r5[j] or "").strip().upper()
            if t in TARGET_MAP:
                tgt[TARGET_MAP[t]] = j
        blocchi_target.append(tgt)

    # tariffe: gruppi in r4 nella zona tariffe, sottoperiodo in r5
    tariffe = {}                      # (gruppo, label_sottop) -> col
    if col_tariffe_start is not None:
        gruppo = None
        for j in range(col_tariffe_start, len(r4)):
            v4 = str(r4[j] or "").strip()
            if v4:
                if "TABELLARE" in v4.upper():
                    gruppo = "tabellare"
                elif "P/U" in v4.upper():
                    gruppo = "pu"
                elif v4.upper() in ("CONTENT",) or (r3[j] and "content" in str(r3[j]).lower()):
                    gruppo = None
            v5 = str(r5[j] or "").strip()
            if gruppo and RE_SOTTOP.match(v5):
                tariffe[(gruppo, v5)] = j
    # content: colonne di r4 dopo l'header r3 'Content'
    col_content = {}
    j_content = next((j for j, v in enumerate(r3)
                      if v and "content" in str(v).lower()), None)
    if j_content is not None:
        for j in range(j_content, len(r4)):
            v = str(r4[j] or "").strip()
            if v and not RE_SOTTOP.match(v) and "30" not in v:
                col_content[j] = v.title()

    # universi (r6)
    universi = {}
    for k, tgt in enumerate(blocchi_target):
        for t, j in tgt.items():
            if r6[j]:
                universi.setdefault(t, int(r6[j]))

    pub, fonte = pubblicato_da_nome(nome)
    if pub is None:
        pub = (wb.properties.modified or wb.properties.created).date()
        fonte = "xlsx_meta"
    # il fallback dai metadati xlsx e' spesso il touch del download: se cade
    # dopo l'inizio del periodo, clampa. Le date STAMPATE invece si tengono
    # anche se tardive (aggiornamenti retroattivi = versioni reali).
    if fonte == "xlsx_meta" and pub > min(d for c in date_blocchi for d in c):
        pub = min(d for c in date_blocchi for d in c)
        fonte = "stimata (clamp a inizio periodo)"

    n_file, prev_righe, rete_cur = 0, [], None
    for riga in righe[7:]:
        if riga[0]:
            rete_cur = str(riga[0]).strip()
        rubrica = str(riga[1] or "").strip()
        if not rubrica:
            continue
        note_r = str(riga[2] or "").strip()
        programma = str(riga[3] or "").strip()
        o1, o2 = orario_split(riga[4])
        mask = "".join("1" if (i in col_giorni and riga[col_giorni[i]]) else "0"
                       for i in range(7))
        content = ",".join(v for j, v in col_content.items()
                           if j < len(riga) and riga[j])
        rete_std = RETE_MAP.get(rete_cur.upper() if rete_cur else "", rete_cur)
        for k, (j0, lab) in enumerate(blocchi):
            d1, d2 = date_blocchi[k]
            val = {}
            for t, j in blocchi_target[k].items():
                v = riga[j] if j < len(riga) else None
                val[t] = round(float(v)) if isinstance(v, (int, float)) else None
            if all(v is None for v in val.values()):
                continue
            t30 = tariffe.get(("tabellare", lab))
            tpu = tariffe.get(("pu", lab))
            righe_csv.append({
                "file": nome, "periodo": periodo_lab, "rete": rete_std,
                "rubrica_vendita": rubrica, "programma": programma,
                "note": note_r,
                "orario_inizio": o1, "orario_fine": o2, "giorni_mask": mask,
                "sottoperiodo": lab, "data_inizio": d1, "data_fine": d2,
                "ind_000": round(val.get("individui", 0) / 1000, 1) if val.get("individui") is not None else "",
                "ra_000": round(val.get("ra", 0) / 1000, 1) if val.get("ra") is not None else "",
                "target_15_64": round(val.get("15_64", 0) / 1000, 1) if val.get("15_64") is not None else "",
                "target_25_54": round(val.get("25_54", 0) / 1000, 1) if val.get("25_54") is not None else "",
                "target_15_34": round(val.get("15_34", 0) / 1000, 1) if val.get("15_34") is not None else "",
                "tariffa_30s": (riga[t30] if t30 is not None and t30 < len(riga) else "") or "",
                "tariffa_pu_30s": (riga[tpu] if tpu is not None and tpu < len(riga) else "") or "",
                "content": content, "aggiornamento": pub.isoformat(),
            })
            n_file += 1
            for t, v in val.items():
                if v is None:
                    continue
                prev_righe.append({
                    "grana": "periodo", "periodo_label": lab,
                    "periodo_da": d1, "periodo_a": d2,
                    "rete": rete_std, "posizione": rubrica,
                    "target": t if t in ("individui", "ra") else t,
                    "target_label": t, "metrica": "amr_migliaia",
                    "valore": v / 1000})
    wb.close()

    doc_id = os.path.splitext(nome)[0].replace(" (1)", "")
    tutte = [d for coppia in date_blocchi for d in coppia]
    conn.execute("DELETE FROM previsione WHERE doc_id = ?", [doc_id])
    conn.execute("DELETE FROM doc_sorgente WHERE doc_id = ?", [doc_id])
    conn.execute("INSERT INTO doc_sorgente VALUES (?,?,?,?,?,?,?,?,?)", [
        doc_id, "rai", "listino_stime", path, min(tutte), max(tutte),
        pub, fonte, f"tabellare video; {len(blocchi)} sottoperiodi; "
        f"universi {universi.get('individui')}"])
    n_prev = previsioni.registra(conn, "rai_listino", doc_id.split("listino_")[-1][:40],
                                 pub, prev_righe, doc_id=doc_id)
    riepilogo.append((nome, periodo_lab, len(blocchi), n_file, n_prev,
                      str(min(tutte)), str(max(tutte)), pub.isoformat()))
    print(f"OK {nome}: {periodo_lab!r} sottoperiodi={len(blocchi)} "
          f"righe_csv={n_file} previsioni={n_prev} "
          f"[{min(tutte)}..{max(tutte)}] pubbl={pub}", flush=True)

conn.close()
campi = list(righe_csv[0].keys())
with open(OUT, "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=campi)
    w.writeheader()
    w.writerows(righe_csv)
print(f"\nCSV: {OUT} — {len(righe_csv)} righe")
